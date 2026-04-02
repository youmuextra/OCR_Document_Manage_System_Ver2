# ui/statistics_dialog.py
"""
统计报表对话框
"""

from PySide6.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QLabel, 
    QPushButton, QTableWidget, QTableWidgetItem,
    QDateEdit, QGroupBox, QFormLayout, QHeaderView, 
    QMessageBox, QFrame, QWidget, QSplitter,
    QComboBox, QLineEdit, QTextEdit, QProgressBar,
    QSizePolicy
)
from PySide6.QtGui import QFont, QPixmap, QPainter, QColor
from PySide6.QtCore import Qt, QDate, QTimer, Signal
from datetime import datetime, timedelta

class StatisticsDialog(QDialog):
    """统计报表对话框"""
    
    def __init__(self, db_manager, current_user, parent=None):
        super().__init__(parent)
        self.db_manager = db_manager
        self.current_user = current_user
        
        # 初始化数据存储
        self.stats_data = None
        self.chart_data = {}
        
        self.setup_ui()
    
    def safe_date_format(self, date_obj, date_format='%Y-%m-%d'):
        """安全地格式化日期，处理字符串和datetime对象"""
        if date_obj is None:
            return "未知"
        
        if hasattr(date_obj, 'strftime'):
            # 如果是datetime对象
            return date_obj.strftime(date_format)
        elif isinstance(date_obj, str):
            # 如果是字符串，尝试转换为datetime再格式化
            from datetime import datetime
            try:
                # 尝试不同的日期格式
                for fmt in ('%Y-%m-%d', '%Y/%m/%d', '%Y-%m-%d %H:%M:%S', '%Y/%m/%d %H:%M:%S'):
                    try:
                        dt = datetime.strptime(date_obj[:10], fmt)
                        return dt.strftime(date_format)
                    except ValueError:
                        continue
                # 如果都无法解析，返回原字符串
                return date_obj[:10]
            except Exception:
                return date_obj[:10]
        else:
            # 其他类型
            return str(date_obj)[:10] if date_obj else "未知"

    def _format_datetime_display(self, value):
        """统一时间显示：去除微秒，仅保留到秒。"""
        if value is None:
            return ""

        if isinstance(value, datetime):
            return value.strftime('%Y-%m-%d %H:%M:%S')

        text = str(value).strip()
        if not text:
            return ""

        text = text.replace('T', ' ')
        # 常见格式：YYYY-MM-DD HH:MM:SS.ffffff -> YYYY-MM-DD HH:MM:SS
        if len(text) >= 19 and text[4] == '-' and text[7] == '-' and text[10] == ' ' and text[13] == ':' and text[16] == ':':
            return text[:19]
        return text

    def _format_date_display(self, value):
        """按日期显示（YYYY-MM-DD）。"""
        if value is None:
            return ""
        if isinstance(value, datetime):
            return value.strftime('%Y-%m-%d')
        text = str(value).strip().replace('T', ' ')
        if not text:
            return ""
        return text[:10] if len(text) >= 10 else text

    def _get_default_chart_color(self, title: str) -> str:
        """按统计主题返回默认图表主色。"""
        title = str(title or '')
        if not any(k in title for k in ('收文', '发文', '流转', '用户活跃度')):
            title = str(self.stat_type_combo.currentText() or title)
        if '收文' in title:
            return '#3B82F6'
        if '发文' in title:
            return '#10B981'
        if '流转' in title:
            return '#F59E0B'
        if '用户活跃度' in title:
            return '#8B5CF6'
        return '#4CAF50'

    def _get_bar_color(self, title: str, category: str, index: int) -> str:
        """按统计主题和条目返回柱状条颜色。"""
        title = str(title or '')
        if not any(k in title for k in ('收文', '发文', '流转', '用户活跃度', '文档数量统计')):
            title = str(self.stat_type_combo.currentText() or title)
        category = str(category or '')

        if '文档数量统计' in title:
            specific_colors = {
                '收文': '#3B82F6',
                '发文': '#10B981',
                '流转': '#F59E0B',
            }
            return specific_colors.get(category, '#4CAF50')

        if '收文' in title:
            palette = ['#3B82F6', '#60A5FA', '#2563EB', '#1D4ED8', '#93C5FD']
            return palette[index % len(palette)]

        if '发文' in title:
            palette = ['#10B981', '#34D399', '#059669', '#047857', '#6EE7B7']
            return palette[index % len(palette)]

        if '流转' in title:
            palette = ['#F59E0B', '#FBBF24', '#D97706', '#B45309', '#FCD34D']
            return palette[index % len(palette)]

        if '用户活跃度' in title:
            palette = ['#8B5CF6', '#A78BFA', '#7C3AED', '#6D28D9', '#C4B5FD']
            return palette[index % len(palette)]

        return self._get_default_chart_color(title)

    def _get_header_background_color(self) -> str:
        """根据当前统计类型返回表头背景色。"""
        return self._get_default_chart_color(self.stat_type_combo.currentText())

    def _get_header_foreground_color(self) -> str:
        """根据表头背景色返回合适的前景色。"""
        stat_type = str(self.stat_type_combo.currentText() or '')
        if '用户活跃度' in stat_type or '流转' in stat_type:
            return '#FFFFFF'
        if '收文' in stat_type or '发文' in stat_type:
            return '#FFFFFF'
        return '#FFFFFF'

    def _get_section_highlight_color(self) -> str:
        """分组小节行使用当前主题的浅色高亮。"""
        stat_type = str(self.stat_type_combo.currentText() or '')
        if '收文' in stat_type:
            return '#DBEAFE'
        if '发文' in stat_type:
            return '#D1FAE5'
        if '流转' in stat_type:
            return '#FEF3C7'
        if '用户活跃度' in stat_type:
            return '#EDE9FE'
        return '#E5E7EB'

    def _get_section_text_color(self) -> str:
        """分组小节行的文字颜色。"""
        stat_type = str(self.stat_type_combo.currentText() or '')
        if '收文' in stat_type:
            return '#1D4ED8'
        if '发文' in stat_type:
            return '#047857'
        if '流转' in stat_type:
            return '#B45309'
        if '用户活跃度' in stat_type:
            return '#6D28D9'
        return '#374151'

    def _apply_time_header_tooltips(self, headers):
        """为时间列添加表头悬停说明，降低理解成本。"""
        tips = {
            '发文时间': '该公文在发文模块登记/发送的时间（精确到秒）。',
            '收文时间': '该公文在收文模块登记接收的时间（精确到秒）。',
            '发起流转时间': '该公文发起流转记录的时间（精确到秒）。',
            '取件时间': '取件人实际取走公文的时间（精确到秒）。',
            '归还时间': '取件/借阅后归还公文的时间（精确到秒）。',
        }
        for idx, h in enumerate(headers or []):
            item = self.table_widget.horizontalHeaderItem(idx)
            if item and h in tips:
                item.setToolTip(tips[h])
    
    def setup_ui(self):
        """设置UI"""
        self.setWindowTitle("统计报表")
        self.resize(1000, 700)
        
        # 主布局
        main_layout = QVBoxLayout()
        main_layout.setSpacing(10)
        
        # 标题
        title_label = QLabel("公文管理系统 - 统计报表")
        title_label.setAlignment(Qt.AlignCenter)
        title_label.setFont(QFont("Microsoft YaHei", 16, QFont.Bold))
        title_label.setStyleSheet("color: #2c3e50; margin: 10px 0;")
        main_layout.addWidget(title_label)
        
        # 查询条件区域
        query_group = QGroupBox("查询条件")
        query_layout = QFormLayout()
        query_layout.setSpacing(10)
        
        # 统计类型
        self.stat_type_combo = QComboBox()
        self.stat_type_combo.addItems([
            "综合明细统计"
        ])
        query_layout.addRow("统计类型:", self.stat_type_combo)
        
        # 时间范围
        time_layout = QHBoxLayout()
        
        self.start_date_edit = QDateEdit()
        self.start_date_edit.setDate(QDate.currentDate().addMonths(-1))
        self.start_date_edit.setCalendarPopup(True)
        
        self.end_date_edit = QDateEdit()
        self.end_date_edit.setDate(QDate.currentDate())
        self.end_date_edit.setCalendarPopup(True)
        
        time_layout.addWidget(QLabel("从"))
        time_layout.addWidget(self.start_date_edit)
        time_layout.addWidget(QLabel("到"))
        time_layout.addWidget(self.end_date_edit)
        time_layout.addStretch()
        
        query_layout.addRow("时间范围:", time_layout)
        
        # 快速选择按钮
        quick_buttons_layout = QHBoxLayout()
        
        for days, label in [(7, "最近7天"), (30, "最近30天"), (90, "最近90天"), (365, "最近一年")]:
            btn = QPushButton(label)
            btn.clicked.connect(lambda _checked=False, d=days: self.set_date_range(d))
            btn.setMaximumWidth(80)
            quick_buttons_layout.addWidget(btn)
        
        quick_buttons_layout.addStretch()
        query_layout.addRow("快速选择:", quick_buttons_layout)
        
        query_group.setLayout(query_layout)
        main_layout.addWidget(query_group)
        
        # 操作按钮
        button_layout = QHBoxLayout()
        
        self.generate_button = QPushButton("生成报表")
        self.generate_button.setStyleSheet("""
            QPushButton {
                background-color: #4CAF50;
                color: white;
                border: none;
                padding: 8px 20px;
                border-radius: 4px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #45a049;
            }
        """)
        self.generate_button.clicked.connect(self.generate_statistics)
        
        self.export_button = QPushButton("导出报表")
        self.export_button.setStyleSheet("""
            QPushButton {
                background-color: #2196F3;
                color: white;
                border: none;
                padding: 8px 20px;
                border-radius: 4px;
            }
            QPushButton:hover {
                background-color: #0b7dda;
            }
        """)
        self.export_button.clicked.connect(self.export_report)
        self.export_button.setEnabled(False)
        
        self.print_button = QPushButton("打印")
        self.print_button.clicked.connect(self.print_report)
        
        self.close_button = QPushButton("关闭")
        self.close_button.clicked.connect(self.reject)
        
        button_layout.addWidget(self.generate_button)
        button_layout.addWidget(self.export_button)
        button_layout.addWidget(self.print_button)
        button_layout.addStretch()
        button_layout.addWidget(self.close_button)
        
        main_layout.addLayout(button_layout)
        
        # 创建分割器用于图表和表格
        self.splitter = QSplitter(Qt.Vertical)
        
        # 图表区域
        self.chart_widget = QWidget()
        chart_layout = QVBoxLayout(self.chart_widget)
        
        # 图表容器
        self.chart_container = QWidget()
        self.chart_container.setMinimumHeight(160)
        self.chart_container.setLayout(QVBoxLayout())
        chart_layout.addWidget(self.chart_container)
        
        # 图表标题
        self.chart_title = QLabel("请先生成统计报表")
        self.chart_title.setAlignment(Qt.AlignCenter)
        self.chart_title.setFont(QFont("Microsoft YaHei", 12, QFont.Bold))
        self.chart_title.setStyleSheet("color: #555; margin: 10px;")
        chart_layout.addWidget(self.chart_title)
        
        self.splitter.addWidget(self.chart_widget)
        
        # 表格区域
        self.table_widget = QTableWidget()
        self.table_widget.setAlternatingRowColors(True)
        self.table_widget.setStyleSheet("""
            QTableWidget {
                alternate-background-color: #f9f9f9;
                selection-background-color: #e3f2fd;
            }
            QTableWidget::item {
                padding: 5px;
            }
        """)
        
        self.splitter.addWidget(self.table_widget)
        
        # 设置分割比例
        self.splitter.setSizes([220, 480])
        
        main_layout.addWidget(self.splitter)
        
        # 状态栏
        self.status_label = QLabel("就绪")
        self.status_label.setStyleSheet("color: #666; font-size: 12px;")
        main_layout.addWidget(self.status_label)
        
        self.setLayout(main_layout)
    
    def set_date_range(self, days):
        """设置日期范围"""
        end_date = QDate.currentDate()
        start_date = end_date.addDays(-days)
        
        self.start_date_edit.setDate(start_date)
        self.end_date_edit.setDate(end_date)
    
    def generate_statistics(self):
        """生成统计报表"""
        # 获取查询条件
        stat_type = self.stat_type_combo.currentText()
        start_date = self.start_date_edit.date().toPython()
        end_date = self.end_date_edit.date().toPython()
        
        # 验证日期范围
        if start_date > end_date:
            QMessageBox.warning(self, "警告", "开始日期不能晚于结束日期")
            return
        
        # 更新状态
        self.status_label.setText("正在生成统计报表...")
        self.generate_button.setEnabled(False)
        
        # 使用QTimer延迟执行，避免界面卡顿
        QTimer.singleShot(100, lambda: self._do_generate_statistics(stat_type, start_date, end_date))
    
    def _do_generate_statistics(self, stat_type, start_date, end_date):
        """实际生成统计报表"""
        try:
            success, message, records = self.db_manager.get_statistics_detail_records(start_date, end_date)
            if not success:
                raise Exception(message)
            self.display_detail_statistics(records, start_date, end_date)
            
            # 启用导出按钮
            self.export_button.setEnabled(True)
            self.status_label.setText(f"统计报表生成完成 - {stat_type}")
            
        except Exception as e:
            QMessageBox.critical(self, "错误", f"生成统计报表失败: {e}")
            self.status_label.setText("统计失败")
        finally:
            self.generate_button.setEnabled(True)

    def display_detail_statistics(self, records, start_date, end_date):
        """显示统一统计明细表（无柱状图）。"""
        start_str = self.safe_date_format(start_date)
        end_str = self.safe_date_format(end_date)
        self.chart_title.setText(f"综合明细统计（{start_str} 至 {end_str}）")
        # 明细模式下仅保留表格，隐藏中间图表/提示区域
        self.chart_widget.hide()
        self.splitter.setSizes([0, 700])

        headers = [
            "文号", "标题", "密级", "紧急程度", "发文单位", "发文时间", "去向/发往", "经办人",
            "收文时间", "发起流转时间", "流转类型", "取件单位", "取件人", "取件时间", "归还时间",
            "状态", "备注"
        ]
        table_data = [headers]

        for r in records or []:
            table_data.append([
                str(r.get('文号', '')),
                str(r.get('标题', '')),
                str(r.get('密级', '')),
                str(r.get('紧急程度', '')),
                str(r.get('发文单位', '')),
                self._format_date_display(r.get('发文时间', '')),
                str(r.get('去向/发往', '')),
                str(r.get('经办人', '')),
                self._format_date_display(r.get('收文时间', '')),
                self._format_datetime_display(r.get('发起流转时间', '')),
                str(r.get('流转类型', '')),
                str(r.get('取件单位', '')),
                str(r.get('取件人', '')),
                self._format_datetime_display(r.get('取件时间', '')),
                self._format_datetime_display(r.get('归还时间', '')),
                str(r.get('状态', '')),
                str(r.get('备注', '')),
            ])

        self.display_table(table_data)
        self.stats_data = {
            'records': records,
            'start_date': start_date,
            'end_date': end_date
        }
    
    def display_overall_statistics(self, data):
        """显示整体统计"""
        try:
            # 使用安全的日期格式化
            start_str = self.safe_date_format(data.get('start_date'))
            end_str = self.safe_date_format(data.get('end_date'))
            
            # 更新图表标题
            self.chart_title.setText(f"整体统计 ({start_str} 至 {end_str})")
            
            # 准备图表数据
            categories = ['收文', '发文', '流转']
            values = [
                data.get('receive_count', 0) or 0,
                data.get('send_count', 0) or 0, 
                data.get('circulation_count', 0) or 0
            ]
        
            # ✅ 修复：确保所有值都是数字
            values = [float(v) if v is not None else 0.0 for v in values]
            # 创建柱状图
            self.create_bar_chart(categories, values, "数量", "文档类型", "文档数量统计")
            
            # 准备表格数据
            table_data = [
                ["统计项", "数量", "占比"],
                ["收文数量", str(data.get('receive_count', 0)), f"{self.calculate_percentage(data.get('receive_count', 0), sum(values))}%"],
                ["发文数量", str(data.get('send_count', 0)), f"{self.calculate_percentage(data.get('send_count', 0), sum(values))}%"],
                ["流转数量", str(data.get('circulation_count', 0)), f"{self.calculate_percentage(data.get('circulation_count', 0), sum(values))}%"],
                ["总计", str(sum(values)), "100%"]
            ]
            
            # 添加流转状态统计
            circulation_status = data.get('circulation_status', {})
            if circulation_status:
                table_data.append(["", "", ""])
                table_data.append(["流转状态", "数量", "占比"])
                
                total_circulation = sum(circulation_status.values())
                for status, count in circulation_status.items():
                    percentage = self.calculate_percentage(count, total_circulation) if total_circulation > 0 else 0
                    table_data.append([status, str(count), f"{percentage}%"])
            
            self.display_table(table_data)
            
            # 保存统计数据
            self.stats_data = data
            self.chart_data = {
                'categories': categories,
                'values': values,
                'title': '整体统计'
            }
            
        except Exception as e:
            print(f"❌ 显示整体统计失败: {e}")
            import traceback
            traceback.print_exc()
            QMessageBox.warning(self, "错误", f"显示统计结果失败: {str(e)}")
    
    def display_send_statistics(self, documents, start_date, end_date):
        """显示发文统计"""
        # 按部门统计发文数量
        dept_stats = {}
        status_stats = {}
        
        for doc in documents:
            dept = doc.get('issuing_unit', '未指定部门')
            dept_stats[dept] = dept_stats.get(dept, 0) + 1
            
            status = doc.get('send_status', '已发文')
            status_stats[status] = status_stats.get(status, 0) + 1
        
        # 图表标题
        start_str = self.safe_date_format(start_date)
        end_str = self.safe_date_format(end_date)
        self.chart_title.setText(f"发文统计 ({start_str} 至 {end_str})")
        
        # 创建部门发文柱状图
        if dept_stats:
            categories = list(dept_stats.keys())
            values = list(dept_stats.values())
            
            # 限制显示的部门数量
            if len(categories) > 10:
                # 只显示前10个部门，其余合并为"其他"
                sorted_items = sorted(zip(values, categories), reverse=True)
                top_values, top_categories = zip(*sorted_items[:10])
                
                other_count = sum(values) - sum(top_values)
                if other_count > 0:
                    categories = list(top_categories) + ['其他']
                    values = list(top_values) + [other_count]
                else:
                    categories = list(top_categories)
                    values = list(top_values)
            
            self.create_bar_chart(categories, values, "数量", "发文单位", "按发文单位统计")
        
        # 准备表格数据
        table_data = [
            ["发文单位", "发文数量", "占比"],
        ]
        
        total_send = len(documents)
        sorted_depts = sorted(dept_stats.items(), key=lambda x: x[1], reverse=True)
        
        for dept, count in sorted_depts:
            percentage = self.calculate_percentage(count, total_send)
            table_data.append([dept, str(count), f"{percentage}%"])
        
        # 添加状态统计
        table_data.append(["", "", ""])
        table_data.append(["发文状态", "数量", "占比"])
        
        for status, count in sorted(status_stats.items(), key=lambda item: item[1], reverse=True):
            if count > 0:
                percentage = self.calculate_percentage(count, total_send)
                table_data.append([status, str(count), f"{percentage}%"])
        
        self.display_table(table_data)
        
        # 保存统计数据
        self.stats_data = {
            'documents': documents,
            'dept_stats': dept_stats,
            'status_stats': status_stats
        }
    
    def display_receive_statistics(self, documents, start_date, end_date):
        """显示收文统计"""
        # 按密级统计
        security_stats = {'普通': 0, '秘密': 0, '机密': 0, '绝密': 0}
        # 按紧急程度统计
        urgency_stats = {'普通': 0, '平急': 0, '加急': 0, '特急': 0, '急件': 0}
        
        for doc in documents:
            security = doc.get('security_level', '普通')
            security_stats[security] = security_stats.get(security, 0) + 1
            
            urgency = doc.get('urgency_level', '普通')
            urgency_stats[urgency] = urgency_stats.get(urgency, 0) + 1
        
        # 图表标题
        start_str = self.safe_date_format(start_date)
        end_str = self.safe_date_format(end_date)
        self.chart_title.setText(f"收文统计 ({start_str} 至 {end_str})")
        
        # 图表展示（稳定版：柱状图，避免PieChart相关兼容问题）
        security_pairs = [(k, v) for k, v in security_stats.items() if v > 0]
        if security_pairs:
            sec_categories = [k for k, _ in security_pairs]
            sec_values = [v for _, v in security_pairs]
            self.create_bar_chart(sec_categories, sec_values, "数量", "密级", "按密级统计")
        else:
            self.create_simple_chart(["普通"], [0], "按密级统计")
        
        # 准备表格数据
        table_data = [
            ["密级", "数量", "占比"],
        ]
        
        total_receive = len(documents)
        
        for security, count in security_stats.items():
            if count > 0:
                percentage = self.calculate_percentage(count, total_receive)
                table_data.append([security, str(count), f"{percentage}%"])
        
        table_data.append(["", "", ""])
        table_data.append(["紧急程度", "数量", "占比"])
        
        for urgency, count in urgency_stats.items():
            if count > 0:
                percentage = self.calculate_percentage(count, total_receive)
                table_data.append([urgency, str(count), f"{percentage}%"])
        
        self.display_table(table_data)
        
        # 保存统计数据
        self.stats_data = {
            'documents': documents,
            'security_stats': security_stats,
            'urgency_stats': urgency_stats
        }
    
    def display_circulation_statistics(self, records, start_date, end_date):
        """显示流转统计"""
        # 按流转类型统计
        type_stats = {'交接': 0, '借阅': 0, '其他': 0}
        # 按状态统计
        status_stats = {'待确认': 0, '流转中': 0, '已借出': 0, '已归还': 0, '已完成': 0}
        
        for record in records:
            circ_type = record.get('circulation_type', '其他')
            type_stats[circ_type] = type_stats.get(circ_type, 0) + 1
            
            status = record.get('status', '待确认')
            status_stats[status] = status_stats.get(status, 0) + 1
        
        # 图表标题
        start_str = self.safe_date_format(start_date)
        end_str = self.safe_date_format(end_date)
        self.chart_title.setText(f"流转统计 ({start_str} 至 {end_str})")
        
        # 创建流转类型柱状图
        categories = list(type_stats.keys())
        values = list(type_stats.values())
        
        self.create_bar_chart(categories, values, "数量", "流转类型", "按流转类型统计")
        
        # 准备表格数据
        table_data = [
            ["流转类型", "数量", "占比"],
        ]
        
        total_circulation = len(records)
        
        for circ_type, count in type_stats.items():
            if count > 0:
                percentage = self.calculate_percentage(count, total_circulation)
                table_data.append([circ_type, str(count), f"{percentage}%"])
        
        table_data.append(["", "", ""])
        table_data.append(["流转状态", "数量", "占比"])
        
        for status, count in status_stats.items():
            if count > 0:
                percentage = self.calculate_percentage(count, total_circulation)
                table_data.append([status, str(count), f"{percentage}%"])
        
        # 添加借阅统计
        borrow_records = [r for r in records if r.get('circulation_type') == '借阅']
        if borrow_records:
            table_data.append(["", "", ""])
            table_data.append(["借阅统计", "数量", "说明"])
            
            overdue_count = 0
            returned_count = len([r for r in borrow_records if r.get('status') == '已归还'])
            borrowed_count = len([r for r in borrow_records if r.get('status') == '已借出'])
            
            # 计算逾期数量
            for record in borrow_records:
                if record.get('status') == '已借出' and record.get('due_date'):
                    due_date = record.get('due_date')
                    if isinstance(due_date, str):
                        try:
                            due_date = datetime.strptime(due_date, "%Y-%m-%d")
                        except:
                            continue
                    
                    if due_date and hasattr(due_date, 'date') and due_date.date() < datetime.now().date():
                        overdue_count += 1
            
            table_data.append(["已归还", str(returned_count), ""])
            table_data.append(["借阅中", str(borrowed_count), ""])
            table_data.append(["逾期", str(overdue_count), f"占比: {self.calculate_percentage(overdue_count, borrowed_count) if borrowed_count > 0 else 0}%"])
        
        self.display_table(table_data)
        
        # 保存统计数据
        self.stats_data = {
            'records': records,
            'type_stats': type_stats,
            'status_stats': status_stats
        }
    
    def display_user_activity(self, start_date, end_date):
        """显示用户活跃度"""
        # 获取用户活跃度数据
        try:
            start_dt = datetime.combine(start_date, datetime.min.time())
            end_dt = datetime.combine(end_date, datetime.max.time())

            with self.db_manager.get_connection() as conn:
                cursor = conn.cursor()
                cursor.execute("""
                    SELECT
                        u.username AS username,
                        u.real_name AS real_name,
                        COUNT(dl.id) AS activity_count
                    FROM users u
                    LEFT JOIN document_logs dl
                        ON dl.user_id = u.id
                        AND dl.created_at BETWEEN ? AND ?
                    GROUP BY u.id, u.username, u.real_name
                    ORDER BY activity_count DESC, u.username ASC
                """, (start_dt, end_dt))
                rows = cursor.fetchall()

            # 转换为字典
            activity_data = []
            for row in rows or []:
                activity_data.append({
                    'username': row['username'],
                    'real_name': row['real_name'],
                    'activity_count': int(row['activity_count'] or 0)
                })

            # 图表标题
            start_str = self.safe_date_format(start_date)
            end_str = self.safe_date_format(end_date)
            self.chart_title.setText(f"用户活跃度统计 ({start_str} 至 {end_str})")

            # 创建用户活跃度柱状图
            if activity_data:
                top_data = activity_data[:10]
                usernames = [f"{data['real_name'] or data['username']}" for data in top_data]
                counts = [data['activity_count'] for data in top_data]
                self.create_bar_chart(usernames, counts, "活跃度", "用户", "用户活跃度统计")
            else:
                self.create_simple_chart(["暂无数据"], [0], "用户活跃度统计")

            # 准备表格数据
            table_data = [
                ["用户名", "真实姓名", "活跃度", "排名"],
            ]

            for i, data in enumerate(activity_data, 1):
                table_data.append([
                    data['username'],
                    data['real_name'] or "",
                    str(data['activity_count']),
                    str(i)
                ])

            self.display_table(table_data)

            # 保存统计数据
            self.stats_data = {
                'activity_data': activity_data,
                'start_date': start_date,
                'end_date': end_date
            }
                
        except Exception as e:
            QMessageBox.warning(self, "查询失败", f"获取用户活跃度失败: {e}")
    
    def create_bar_chart(self, categories, values, ylabel, xlabel, title):
        """创建柱状图"""
        try:
            layout = self.chart_container.layout()
            if layout is None:
                layout = QVBoxLayout(self.chart_container)
                self.chart_container.setLayout(layout)

            while layout.count():
                item = layout.takeAt(0)
                widget = item.widget()
                child_layout = item.layout()
                if widget:
                    widget.deleteLater()
                elif child_layout:
                    while child_layout.count():
                        sub_item = child_layout.takeAt(0)
                        sub_widget = sub_item.widget()
                        if sub_widget:
                            sub_widget.deleteLater()

            chart_title_label = QLabel(title)
            chart_title_label.setAlignment(Qt.AlignCenter)
            chart_title_label.setStyleSheet(
                f"font-weight: bold; color: {self._get_default_chart_color(title)}; margin: 4px 0 8px 0;"
            )
            layout.addWidget(chart_title_label)

            if not categories or not values:
                empty_label = QLabel("暂无图表数据")
                empty_label.setAlignment(Qt.AlignCenter)
                empty_label.setStyleSheet("color: #888; padding: 12px;")
                layout.addWidget(empty_label)
                layout.addStretch()
                return

            max_value = max([float(v) for v in values] or [0.0])
            max_value = max(max_value, 1.0)

            for index, (category, value) in enumerate(zip(categories, values)):
                bar_color = self._get_bar_color(title, category, index)
                row_widget = QWidget()
                row_layout = QHBoxLayout(row_widget)
                row_layout.setContentsMargins(8, 2, 8, 2)
                row_layout.setSpacing(8)

                label = QLabel(str(category))
                label.setFixedWidth(130)
                label.setAlignment(Qt.AlignRight | Qt.AlignVCenter)

                value_label = QLabel(str(int(value) if float(value).is_integer() else round(float(value), 2)))
                value_label.setFixedWidth(50)
                value_label.setAlignment(Qt.AlignRight | Qt.AlignVCenter)

                bar = QProgressBar()
                bar.setRange(0, 1000)
                bar.setValue(int((float(value) / max_value) * 1000))
                bar.setTextVisible(False)
                bar.setMinimumHeight(18)
                bar.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
                bar.setStyleSheet("""
                    QProgressBar {
                        border: 1px solid #d0d7de;
                        border-radius: 4px;
                        background-color: #f3f4f6;
                    }
                    QProgressBar::chunk {
                        background-color: %s;
                        border-radius: 4px;
                    }
                """ % bar_color)

                row_layout.addWidget(label)
                row_layout.addWidget(value_label)
                row_layout.addWidget(bar, 1)
                layout.addWidget(row_widget)

            layout.addStretch()
        except Exception as e:
            print(f"创建柱状图失败: {e}")
            self.create_simple_chart(categories, values, title)
    
    def create_simple_chart(self, categories=None, values=None, title="统计图表"):
        """创建简单的文本图表（当没有pyqtgraph时）"""
        try:
            # ✅ 修复：检查布局是否存在
            layout = self.chart_container.layout()
            if layout is None:
                # 如果没有布局，创建一个
                layout = QVBoxLayout(self.chart_container)
                self.chart_container.setLayout(layout)
            
            # 清除之前的图表
            for i in reversed(range(layout.count())):
                widget = layout.itemAt(i).widget()
                if widget:
                    widget.deleteLater()
            
            # 创建简单的文本显示
            text_widget = QTextEdit()
            text_widget.setReadOnly(True)
            text_widget.setFont(QFont("Consolas", 10))
            
            if categories and values:
                # 生成简单的ASCII图表
                chart_text = f"{title}\n"
                chart_text += "=" * 40 + "\n\n"
                
                max_value = max(values) if values else 1
                max_bar_length = 50
                
                for i, (category, value) in enumerate(zip(categories, values)):
                    bar_length = int((value / max_value) * max_bar_length) if max_value > 0 else 0
                    bar = "█" * bar_length
                    chart_text += f"{category[:15]:15} {value:5} {bar}\n"
                
                text_widget.setText(chart_text)
            
            layout.addWidget(text_widget)
            
        except Exception as e:
            print(f"创建简单图表失败: {e}")
            # 如果还是失败，使用最简化的方法
            self.create_minimal_chart(title)
    
    def create_minimal_chart(self, title="统计图表"):
        """创建最小化的图表（最后的备选方案）"""
        try:
            # 确保chart_container有布局
            if self.chart_container.layout() is None:
                self.chart_container.setLayout(QVBoxLayout())
            
            layout = self.chart_container.layout()
            
            # 清除之前的图表
            for i in reversed(range(layout.count())):
                widget = layout.itemAt(i).widget()
                if widget:
                    widget.deleteLater()
            
            # 创建一个简单的标签显示标题
            label = QLabel(f"统计图表: {title}\n\n(图表功能暂时不可用)")
            label.setAlignment(Qt.AlignCenter)
            label.setFont(QFont("Microsoft YaHei", 12))
            label.setStyleSheet("color: #666; padding: 20px;")
            
            layout.addWidget(label)
            
        except Exception as e:
            print(f"创建最小化图表失败: {e}")
    
    def display_table(self, data):
        """显示表格数据"""
        if not data:
            self.table_widget.setRowCount(0)
            self.table_widget.setColumnCount(0)
            return
        
        # data[0] 是表头，其余为数据行
        rows = len(data)
        cols = len(data[0]) if data else 0
        data_rows = max(rows - 1, 0)

        self.table_widget.setRowCount(data_rows)
        self.table_widget.setColumnCount(cols)
        
        # 设置表头
        if rows > 0:
            self.table_widget.setHorizontalHeaderLabels(data[0])
            self._apply_time_header_tooltips(data[0])
            header_bg = self._get_header_background_color()
            header_fg = self._get_header_foreground_color()
            self.table_widget.horizontalHeader().setStyleSheet(
                f"QHeaderView::section {{ background-color: {header_bg}; color: {header_fg}; padding: 6px; border: none; font-weight: bold; }}"
            )
        
        # 填充数据
        section_bg = QColor(self._get_section_highlight_color())
        section_fg = QColor(self._get_section_text_color())
        spacer_bg = QColor('#F9FAFB')
        headers = [str(h).strip() for h in (data[0] or [])]

        for row in range(1, rows):
            row_values = data[row] if row < len(data) else []
            non_empty_values = [str(v).strip() for v in row_values if str(v).strip()]
            is_spacer_row = len(non_empty_values) == 0
            is_section_row = (not is_spacer_row and len(non_empty_values) >= 2 and non_empty_values[1] == '数量')

            for col in range(cols):
                if row < len(data) and col < len(data[row]):
                    item = QTableWidgetItem(str(data[row][col]))

                    header_text = headers[col] if col < len(headers) else ''
                    if col == 0:
                        item.setTextAlignment(Qt.AlignLeft | Qt.AlignVCenter)
                    elif header_text in ('数量', '占比', '活跃度', '排名'):
                        item.setTextAlignment(Qt.AlignCenter)
                    else:
                        item.setTextAlignment(Qt.AlignLeft | Qt.AlignVCenter)

                    if is_section_row:
                        item.setFont(QFont("Microsoft YaHei", 10, QFont.Bold))
                        item.setBackground(section_bg)
                        item.setForeground(section_fg)
                    elif is_spacer_row:
                        item.setBackground(spacer_bg)
                        item.setText('')

                    self.table_widget.setItem(row - 1, col, item)
        
        # 调整列宽
        self.table_widget.resizeColumnsToContents()
        self.table_widget.verticalHeader().setDefaultSectionSize(26)

        header = self.table_widget.horizontalHeader()
        header.setStretchLastSection(False)

        for col, header_text in enumerate(headers):
            if col == 0:
                header.setSectionResizeMode(col, QHeaderView.Stretch)
            elif header_text in ('发文时间', '收文时间', '发起流转时间', '取件时间', '归还时间'):
                header.setSectionResizeMode(col, QHeaderView.Fixed)
                self.table_widget.setColumnWidth(col, 170)
            elif header_text == '状态':
                header.setSectionResizeMode(col, QHeaderView.Fixed)
                self.table_widget.setColumnWidth(col, 220)
            elif header_text == '数量':
                header.setSectionResizeMode(col, QHeaderView.Fixed)
                self.table_widget.setColumnWidth(col, 70)
            elif header_text == '占比':
                header.setSectionResizeMode(col, QHeaderView.Fixed)
                self.table_widget.setColumnWidth(col, 90)
            elif header_text in ('活跃度', '排名'):
                header.setSectionResizeMode(col, QHeaderView.Fixed)
                self.table_widget.setColumnWidth(col, 80)
            else:
                header.setSectionResizeMode(col, QHeaderView.ResizeToContents)
    
    def calculate_percentage(self, part, total):
        """计算百分比"""
        if total == 0:
            return 0
        return round((part / total) * 100, 2)
    
    def export_report(self):
        """导出报表"""
        from PySide6.QtWidgets import QFileDialog
        
        # 选择保存路径
        file_name, _ = QFileDialog.getSaveFileName(
            self,
            "导出报表",
            f"公文统计报表_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            "CSV文件 (*.csv);;Excel文件 (*.xlsx);;文本文件 (*.txt)"
        )
        
        if not file_name:
            return
        
        try:
            if file_name.endswith('.csv'):
                self.export_to_csv(file_name)
            elif file_name.endswith('.xlsx'):
                self.export_to_xlsx(file_name)
            elif file_name.endswith('.txt'):
                self.export_to_txt(file_name)
            else:
                QMessageBox.warning(self, "格式不支持", "目前支持CSV/Excel/TXT格式导出")
                return
            
            QMessageBox.information(self, "成功", f"报表已导出到: {file_name}")
            
        except Exception as e:
            QMessageBox.critical(self, "导出失败", f"导出报表失败: {e}")
    
    def export_to_csv(self, file_name):
        """导出为CSV"""
        import csv
        
        with open(file_name, 'w', newline='', encoding='utf-8-sig') as csvfile:
            writer = csv.writer(csvfile)
            
            # 获取表格数据
            rows = self.table_widget.rowCount()
            cols = self.table_widget.columnCount()

            # 写入表头
            headers = []
            for col in range(cols):
                item = self.table_widget.horizontalHeaderItem(col)
                headers.append(item.text() if item else "")
            writer.writerow(headers)

            time_headers = {'发文时间', '收文时间', '发起流转时间', '取件时间', '归还时间'}
            time_col_indexes = {idx for idx, h in enumerate(headers) if h in time_headers}
            
            for row in range(rows):
                row_data = []
                for col in range(cols):
                    item = self.table_widget.item(row, col)
                    text = item.text() if item else ""
                    # 避免Excel打开CSV时把时间转成数值导致显示#####
                    if col in time_col_indexes and text:
                        text = f"\t{text}"
                    row_data.append(text)
                writer.writerow(row_data)

    def export_to_xlsx(self, file_name):
        """导出为Excel。"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = "统计明细"

        # 表头
        headers = []
        for col in range(self.table_widget.columnCount()):
            item = self.table_widget.horizontalHeaderItem(col)
            headers.append(item.text() if item else "")
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.freeze_panes = "A2"
        if headers:
            ws.auto_filter.ref = f"A1:{chr(64 + len(headers))}1" if len(headers) <= 26 else None

        time_headers = {'发文时间', '收文时间', '发起流转时间', '取件时间', '归还时间'}
        time_col_indexes = [idx + 1 for idx, h in enumerate(headers) if h in time_headers]
        col_max_len = [len(str(h or '')) for h in headers]

        # 数据
        for row in range(self.table_widget.rowCount()):
            row_data = []
            for col in range(self.table_widget.columnCount()):
                item = self.table_widget.item(row, col)
                text = item.text() if item else ""
                row_data.append(text)
                col_max_len[col] = max(col_max_len[col], len(str(text)))
            ws.append(row_data)

        # 时间列强制按文本写入，避免Excel显示#####
        for col_idx in time_col_indexes:
            for row_idx in range(2, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                val = '' if cell.value is None else str(cell.value)
                cell.value = val
                cell.number_format = '@'
                cell.alignment = Alignment(horizontal='left', vertical='center')

        # 自适应列宽（时间列给最小宽度）
        for col_idx, max_len in enumerate(col_max_len, start=1):
            is_time_col = col_idx in time_col_indexes
            target_width = max(max_len + 2, 21 if is_time_col else 10)
            target_width = min(target_width, 60)
            ws.column_dimensions[get_column_letter(col_idx)].width = target_width

        wb.save(file_name)
    
    def export_to_txt(self, file_name):
        """导出为TXT"""
        with open(file_name, 'w', encoding='utf-8') as txtfile:
            # 写入标题
            txtfile.write(f"公文管理系统统计报表\n")
            txtfile.write(f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            txtfile.write(f"统计类型: {self.stat_type_combo.currentText()}\n")
            txtfile.write(f"时间范围: {self.start_date_edit.date().toString('yyyy-MM-dd')} 至 {self.end_date_edit.date().toString('yyyy-MM-dd')}\n")
            txtfile.write("=" * 60 + "\n\n")
            
            # 写入表格数据
            rows = self.table_widget.rowCount()
            cols = self.table_widget.columnCount()

            # 写入表头
            headers = []
            for col in range(cols):
                item = self.table_widget.horizontalHeaderItem(col)
                headers.append(item.text() if item else "")
            txtfile.write(" | ".join(headers) + "\n")
            txtfile.write("-" * 120 + "\n")
            
            for row in range(rows):
                row_data = []
                for col in range(cols):
                    item = self.table_widget.item(row, col)
                    row_data.append(item.text() if item else "")

                txtfile.write(" | ".join(row_data) + "\n")
    
    def print_report(self):
        """打印报表"""
        QMessageBox.information(self, "提示", "打印功能将在后续版本中实现")
